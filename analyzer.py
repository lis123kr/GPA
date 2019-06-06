"""
 Copyright [2017] [Il Seop Lee / Chungbuk National University]
 
 Licensed under the Apache License, Version 2.0 (the "License");
 you may not use this file except in compliance with the License.
 You may obtain a copy of the License at
 
 http://www.apache.org/licenses/LICENSE-2.0
 
 Unless required by applicable law or agreed to in writing, software
 distributed under the License is distributed on an "AS IS" BASIS,
 WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
 See the License for the specific language governing permissions and
 limitations under the License.
 """
from openpyxl import Workbook
import numpy as np
import pandas as pd
from logging import error, info
from datetime import datetime

def infolog(message):
	info("{0} {1}".format(datetime.now().isoformat(), message))
	print("{0} {1}".format(datetime.now().isoformat(), message))

def errorlog(message):
	error("{0} {1}".format(datetime.now().isoformat(), message))	
	print("{0} {1}".format(datetime.now().isoformat(), message))

def next_col(asc):
	asc2 = list(asc)
	if len(asc2) == 1 and asc2[0] != 'Z':
		return chr(ord(asc2[0])+1)
	elif len(asc2) == 1 and asc2[0] == 'Z':
		return 'AA'
	else:
		if asc[-1] != 'Z':
			return asc[:-1] + chr(ord(asc[-1])+1)
		else:
			p = None
			for i in range(-1, -len(asc)-1, -1):
				p = i
				if asc2[i] == 'Z':
					asc2[i] = 'A'
				else:
						break
			if p == -len(asc) and asc[p] == 'Z':
				return 'A'+''.join(asc2)
			else:
				asc = ''.join(asc2)
				return asc[:p] + chr(ord(asc[p])+1) + asc[p+1:]
				
def next_n_col(asc, n_):
	for _ in range(n_):
		asc = next_col(asc)
	return asc

# module in openpyxl v2.5.3 
def expand_levels(levels):
    """
    Multiindexes need expanding so that subtitles repeat
    """
    widths = (len(s) for s in levels)
    widths = list(accumulate(widths, operator.mul))
    size = max(widths)

    for level, width in zip(levels, widths):
        padding = int(size/width) # how wide a title should be
        repeat = int(width/len(level)) # how often a title is repeated
        row = []
        for v in level:
            title = [None]*padding
            title[0] = v
            row.extend(title)
        row = row*repeat
        yield row

# module in openpyxl v2.5.3 
def dataframe_to_rows(df, index=True, header=True):
    """
    Convert a Pandas dataframe into something suitable for passing into a worksheet.
    If index is True then the index will be included, starting one row below the header.
    If header is True then column headers will be included starting one column to the right.
    Formatting should be done by client code.
    """
    import numpy
    from pandas import Timestamp
    blocks = df._data.blocks
    ncols = sum(b.shape[0] for b in blocks)
    data = [None] * ncols

    for b in blocks:
        values = b.values

        if b.dtype.type == numpy.datetime64:
            values = numpy.array([Timestamp(v) for v in values.ravel()])
            values = values.reshape(b.shape)

        result = values.tolist()

        for col_loc, col in zip(b.mgr_locs, result):
            data[col_loc] = col

    if header:
        if df.columns.nlevels > 1:
            rows = expand_levels(df.columns.levels)
        else:
            rows = [list(df.columns.values)]
        for row in rows:
            n = []
            for v in row:
                if isinstance(v, numpy.datetime64):
                    v = Timestamp(v)
                n.append(v)
            row = n
            if index:
                row = [None]*df.index.nlevels + row
            yield row

    cols = None
    if df.index.nlevels > 1:
        cols = zip(*expand_levels(df.index.levels))

    if index:
        yield df.index.names

    for idx, v in enumerate(df.index):
        row = [data[j][idx] for j in range(ncols)]
        if index:
            if cols:
                v = list(next(cols))
            else:
                v = [v]
            row = v + row
        yield row


class Analyzer(object):
	def __init__(self, book=None):
		"""		
			This program was made by Python, PyQt5, Pyinstaller, openpyxl, pandas, numpy.
			If you want to make excutable files(.exe) using Pyinstaller, you must install above modules.

			# Developed by IL Seop Lee, Software Engineering Department in CBNU, 2017	
			# Refactoring, 2018-03		
		"""
		if book is not None:
			self.book = book

			self.book.initialize()
			
			# get raw data of full basepair sequence
			self.book.mergesequence()
			# 뒤로 미룹시다
			# get BP35, major, minor, index of major, minor
			self.book.preprocessing()

			#self.book.save_data(self.book, self.book.filename+'_book.pkl')

	def Extract_difference_of_minor(self, x, y, type_, book):
		# x, y : PxRaw (removed '-')
		if len(x) > len(y):
			y = pd.DataFrame(y, index = x.index)
		else:
			x = pd.DataFrame(x, index = y.index)
        
		tmp_x = x[ x[book.col_DumaPosition] == y[book.col_DumaPosition]]
		tmp_y = y[ x[book.col_DumaPosition] == y[book.col_DumaPosition]]
		x_bp = -tmp_x[book.col_basepair]
		x_bps = x_bp.values.argsort(axis=1)
		x_minor_idx = pd.DataFrame(x_bps[:,1], index=[x_bp.index])
		x_major_idx = pd.DataFrame(x_bps[:,0], index=[x_bp.index])
    
		y_bp = -tmp_y[book.col_basepair]
		y_bps = y_bp.values.argsort(axis=1)
		y_minor_idx = pd.DataFrame(y_bps[:,1], index=[y_bp.index])
		y_major_idx = pd.DataFrame(y_bps[:,0], index=[y_bp.index])
    
		tmp_x = tmp_x[ (x_major_idx == y_major_idx).values.tolist() ]
		tmp_y = tmp_y[ (x_major_idx == y_major_idx).values.tolist() ]
    
		x_major, x_minor = book.get_major_minor(tmp_x[book.col_basepair])
		y_major, y_minor = book.get_major_minor(tmp_y[book.col_basepair])
		x_minor = pd.DataFrame(x_minor.values, index = tmp_x.index)
		y_minor = pd.DataFrame(y_minor.values, index = tmp_y.index)

		x_sum = pd.DataFrame(tmp_x[book.col_basepair].sum(axis=1), index = tmp_x.index)
		y_sum = pd.DataFrame(tmp_y[book.col_basepair].sum(axis=1), index = tmp_y.index)
    
		x_maf = np.divide(x_minor, x_sum) * 100
		y_maf = np.divide(y_minor, y_sum) * 100
		if type_ == 'ASC':
			return tmp_x[((y_maf - x_maf) >= 5.0).values] , x_major.loc[tmp_x.index], x_minor.loc[tmp_x.index]
		else:
			return tmp_x[((x_maf - y_maf) >= 5.0).values] , x_major.loc[tmp_x.index], x_minor.loc[tmp_x.index]

	def Analyze(self, Analyze_type, book):
		try:
			if Analyze_type == "Difference_of_Minor":
				types_ = ["INC", "DEC"]
				from lowhigh import LowHigh
				lh = LowHigh(book, isAverage=False)
				for i in range(0,2):
					infolog("{0} analysis".format(types_[i]))
					wb = Workbook()
					ws1 = wb.active				
					ws3 = wb.create_sheet("Genome structure")
					ws4 = wb.create_sheet("ORF")
					ws5 = wb.create_sheet("NCR")
					ws6 = wb.create_sheet("Base composition_1")
					
					lh.PolymorphicSite(ws1, types_[i], book)
					lh.GenomeStr(ws3,types_[i], book)
					lh.ORFNCR(ws4, types_[i], "ORF", book)
					lh.ORFNCR(ws5, types_[i], "NCR", book)
					lh.BaseComp(ws6, types_[i], book)

					for n_ in range(book.nsheets):
						ws = wb.create_sheet(book.sheet_list[n_])
						lh.listofdata(ws, n_, types_[i], book)

					wb.save("["+types_[i]+"분석]" + book.filename + '.xlsx')
			elif Analyze_type == 'Full': # full
				from fullseq import FullSeq
				fs = FullSeq(book)
				wb = Workbook()
				ws1 = wb.active
				ws2 = wb.create_sheet("GPS")
				ws3 = wb.create_sheet("Genome structure")
				ws4 = wb.create_sheet("ORF")
				ws5 = wb.create_sheet("NCR")
				ws6 = wb.create_sheet("Base composition_1")
				ws7 = wb.create_sheet("Base composition_2")
				ws8 = wb.create_sheet('listOfMaf5')
				# FullSeq.init_full(book)
				
				fs.sheet1(ws1, book)				
				fs.sheet2(ws2, book)			
				fs.sheet3(ws3, book)			
				fs.sheet4_5(ws4, "ORF", book)			
				fs.sheet4_5(ws5, "NCR", book)			
				fs.sheet6(ws6, book)			
				fs.sheet7(ws7, book)
				fs.maf5list(ws8, book)
				wb.save("[분석]" + book.filename + '.xlsx')
			elif Analyze_type == 'Average':
				types_ = ["INC", "DEC"]
				from lowhigh import LowHigh
				lh = LowHigh(book, isAverage=True)
				for i in range(0,2):				
					wb = Workbook()
					ws1 = wb.active				
					ws3 = wb.create_sheet("Genome structure")
					ws4 = wb.create_sheet("ORF")
					ws5 = wb.create_sheet("NCR")
					ws6 = wb.create_sheet("Base composition_1")
					
					lh.PolymorphicSite(ws1, types_[i], book)
					lh.GenomeStr(ws3,types_[i], book)
					lh.ORFNCR(ws4, types_[i], "ORF", book)
					lh.ORFNCR(ws5, types_[i], "NCR", book)
					lh.BaseComp(ws6, types_[i], book)

					wb.save("["+types_[i]+"분석]" + book.filename + '.xlsx')
			elif Analyze_type == 'Cartesian':
				from lowhigh import LowHigh
				lh = LowHigh(book, isCartesian=True)			
				wb = Workbook()
				ws1 = wb.active				
				ws3 = wb.create_sheet("Genome structure")
				ws4 = wb.create_sheet("ORF")
				ws5 = wb.create_sheet("NCR")
				# ws6 = wb.create_sheet("Base composition_1")
				
				lh.PolymorphicSite(ws1, 'INC', book)
				lh.GenomeStr(ws3,'INC', book)
				lh.ORFNCR(ws4, 'INC', "ORF", book)
				lh.ORFNCR(ws5, 'INC', "NCR", book)
				# lh.BaseComp(ws6, 'INC', book)

				wb.save("[Avg_Cts]"+book.filename + '.xlsx')

			# book.P0.to_excel('[통합]'+book.filename+'.xlsx', index=False)

			return "Success"
		except PermissionError:
			return "Permission"
		except AssertionError as e:			
			return e #"Error"
		except Exception as e:
			infolog(e)
			print(e)
			return "Error"





	
