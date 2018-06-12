from analyzer import infolog, next_col #Analyzer.get_Number_of_GPS, next_col, infolog
from numpy import logical_and, logical_or
class LowHigh:
	"""
		analysis of increase or decrease sequence
	"""
	BPmergedinc = []

	BPmergeddec = []

	s1 = [5.0, 5.0, 15.0, 25.0]

	s2 = [51.0, 15.0, 25.0, 51.0]

	def __init__(self, book):
		infolog("lowhigh init start")
		sheet_names = []
		BPmaf = []
		self.dumascol = [book.col_GenomeStructure, book.col_RepeatRegion, book.col_ORF, book.col_DumaPosition, book.col_DumaSeq]
		for bp in book.BP35:
			#get maf 5%
			s, l, minor_ = book.get_Number_of_GPS(bp, 5.0, 51.0)
			BPmaf.append(bp.loc[minor_.index])

		import pandas as pd
		from numpy import divide
		for i, x in enumerate(BPmaf):
			for j, y in enumerate(BPmaf):
				if i >= j:
					continue
				infolog("{0} -> {1}".format(i,j))
				sheet_names.append(book.sheet_list[i]+'->'+book.sheet_list[j])
				# low의 maf5%이상과 high의 maf5%이상 값을 병합
				# 둘 중 하나라도 sum >= 35.0 and maf 5% 이상인 값에서 분석을 진행
				dumaspos = pd.merge(x,y, how='outer', on=self.dumascol, suffixes=('_x', '_y'), right_index=True, left_index=True)[[book.col_DumaPosition]]

				# dumas position 기준으로 양쪽 위치 추출
				x1 = book.BP35[i][ book.BP35[i][book.col_DumaPosition].isin(dumaspos[book.col_DumaPosition].values.tolist())] if len(dumaspos) is not 0 else pd.DataFrame()
				y1 = book.BP35[j][ book.BP35[j][book.col_DumaPosition].isin(dumaspos[book.col_DumaPosition].values.tolist())]	if len(dumaspos) is not 0 else pd.DataFrame()

<<<<<<< HEAD
				merged = pd.merge(x1,y1, how='outer', on=self.dumascol, suffixes=('_x', '_y'), right_index=True, left_index=True)
				print("x1, y1", len(x1), len(y1))
				print("merged2 ", len(merged))
=======
				merged = pd.merge(x1,y1, how='outer', on=dumascol, suffixes=('_x', '_y'), right_index=True, left_index=True)

				# print("x1, y1", len(x1), len(y1))
				# print("merged2 ", len(merged))

>>>>>>> b1e44ac00ee06e43e40cd7d591d20644bb8c0f16
				if(len(x1) is 0 or len(y1) is 0):
					self.BPmergedinc.append(pd.DataFrame())
					self.BPmergeddec.append(pd.DataFrame())
					continue

				# 0일수도 예외처리..
				x_maf = divide(x1[['minor']], x1[['sum']]) * 100
				y_maf = divide(y1[['minor']], y1[['sum']]) * 100
				
				# 증가하는 위치와 감소하는 위치
				## 양쪽에 서로 없는 index가 있으면 Error - 한쪽에 값이 없으면 NaN : dropna함수로 제거
				
				merged['diffofinc'] = y_maf - x_maf
				merged['diffofdec'] = x_maf - y_maf

<<<<<<< HEAD
				# 한쪽에 값이 없으면 drop되기 때문에 어디쪽에 해도 상관은 없음	
				# github 가져오기			
				self.BPmergedinc.append(pd.DataFrame.dropna(merged[(incidx>=5.0).values], how='all'))
				self.BPmergeddec.append(pd.DataFrame.dropna(merged[(decidx>=5.0).values], how='all'))
=======
				# added 0605
				cond1 = merged['major_idx_x'] == merged['major_idx_y']
				cond2 = merged['diffofinc'] >= 5.0
				cond3 = merged['diffofdec'] >= 5.0
				# 한쪽에 값이 없으면 drop되기 때문에 어디쪽에 해도 상관은 없음				
				self.BPmergedinc.append(pd.DataFrame.dropna(merged[ logical_and(cond1, cond2).values ], how='all'))
				self.BPmergeddec.append(pd.DataFrame.dropna(merged[ logical_and(cond1, cond3).values ], how='all'))
>>>>>>> b1e44ac00ee06e43e40cd7d591d20644bb8c0f16

		book.sheet_list = sheet_names
		book.nsheets = len(sheet_names)
		infolog("lowhigh init end")

	def PolymorphicSite(self, ws, types_, book):
		infolog("lowhigh PolymorphicSite start")
		ws.title = "INC-Polymorphic site"
		# A col
		ws['A1'] = types_
		ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
		
		# C col
		ws['C1'] = "Increase in genetic polymrphism (%)"
		ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)

		ws['C2'] = "5≤n<15"
		ws['D2'] = "15≤n<25"
		ws['E2'] = "25≤n"
		ws['F2'] = "Sum"
		from numpy import logical_and
		for i in range(book.nsheets):
			bp = self.BPmergedinc[i][['diffofinc']] if types_ == 'INC' else self.BPmergeddec[i][['diffofdec']]

			ws['A' + str(i+3)] = book.sheet_list[i]
			ws.merge_cells(start_row=i+3, start_column=1, end_row=i+3, end_column=2)
			ws['C' + str(i+3)] = len(bp[ logical_and(bp >= 5.0, bp < 15).values]) if len(bp) is not 0 else 0
			ws['D' + str(i+3)] = len(bp[ logical_and(bp >= 15.0, bp < 25).values]) if len(bp) is not 0 else 0
			ws['E' + str(i+3)] = len(bp[ logical_and(bp >= 25.0, bp < 51).values]) if len(bp) is not 0 else 0
			ws['F' + str(i+3)] = len(bp[ (bp >= 5.0).values ])
		infolog("lowhigh PolymorphicSite end")

	def GenomeStr(self, ws, types_, book):
		infolog("lowhigh GenomeStr start")
		bp = self.BPmergedinc if types_ == 'INC' else self.BPmergeddec		
		col_diff = "diffofinc" if types_ == 'INC' else "diffofdec"

		# s1 ~ s2 
		for i in range(len(self.s1)):
			rows = i*(14+len(book.GenomeStructure) + len(book.RepeatRegion))+ 1

			ws['A'+str(rows)] = 'Region'
			ws.merge_cells(start_row=rows, start_column=1, end_row=rows+1, end_column=1)
			ws['B'+str(rows)] = 'Dumas Length'
			ws.merge_cells(start_row=rows, start_column=2, end_row=rows+1, end_column=2)

			for r in range(2, 2+len(book.GenomeStructure)):
				ws['A' + str(rows+r)] = book.GenomeStructure[r-2]
				ws['B' + str(rows+r)] = len(book.Dumas[ book.Dumas[book.col_GenomeStructure] == book.GenomeStructure[r-2] ])

			ws['A' + str(rows+2+len(book.GenomeStructure))] = 'Total'
			ws['B' + str(rows+2+len(book.GenomeStructure))] = len(book.Dumas)

			sum_ = 0
			for r in range(len(book.RepeatRegion)):
				row = rows+r+3+len(book.GenomeStructure)
				ws['A' + str(row)] = book.RepeatRegion[r]
				cnt = len(book.Dumas[ book.Dumas[book.col_RepeatRegion] == book.RepeatRegion[r] ])
				ws['B' + str(row)] = cnt
				sum_ += cnt

			leng = rows+ len(book.GenomeStructure) + len(book.RepeatRegion)
			ws['A' + str(leng+3)] = 'Total'
			ws['B' + str(leng+3)] = sum_
			ws['A' + str(leng+4)] = 'ORF'
			ws['B' + str(leng+4)] = len(book.Dumas[ book.Dumas[book.col_ORF].isin(book.ORF)])	
			ws['A' + str(leng+5)] = 'NCR'
			ws['B' + str(leng+5)] = len(book.Dumas[ book.Dumas[book.col_ORF].isin(book.NCR)])

			col = 'C'
			ws[col+str(rows)] = '{0} in genetic polymorphism'.format("Increase" if types_ is 'INC' else "Decrease")
			ws.merge_cells(start_row=rows, start_column=3, end_row=rows, end_column=2+book.nsheets)
			ws[chr(ord(col)+book.nsheets)+str(rows)] = "Average {0}".format("Increase" if types_ is 'INC' else "Decrease")
			ws.merge_cells(start_row=rows, start_column=3+book.nsheets, end_row=rows, end_column=2+2*book.nsheets)
			for s in range(book.nsheets):
				idx = logical_and( bp[s][col_diff] >= self.s1[i], bp[s][col_diff] < self.s2[i] )
				bpx = bp[s][idx.values]
				
				ncol = chr(ord(col)+book.nsheets)
				ws[col+str(rows+1)] = book.sheet_list[s]

				sum_, avg_ = 0, 0
				for r in range(2, 2+len(book.GenomeStructure)):
					cnt = bpx[ bpx[book.col_GenomeStructure] == book.GenomeStructure[r-2]][[col_diff]]
					ws[col+str(rows+r)] = len(cnt)
					ws[ncol+str(rows+r)] = str(round((cnt.sum()[0] / len(cnt)), 3))+'%' if len(cnt) is not 0 else 'N/A'
					avg_ += cnt.sum()[0]
					sum_ += len(cnt)
				# Total
				ws[col+str(rows+2+len(book.GenomeStructure))] = sum_
				ws[chr(ord(col)+book.nsheets)+str(rows+2+len(book.GenomeStructure))] = str(round(avg_/sum_, 3))+'%' if sum_ is not 0 else 'N/A'

				sum_, avg_ = 0, 0
				for r in range(len(book.RepeatRegion)):
					row = rows+r+3+len(book.GenomeStructure)
					cnt = bpx[ bpx[book.col_RepeatRegion] == book.RepeatRegion[r]][[col_diff]]
					ws[col+str(row)] = len(cnt)
					ws[ncol+str(row)] = str(round((cnt.sum()[0] / len(cnt)),3)) if len(cnt) is not 0 else 'N/A'
					sum_ += len(cnt)
					avg_ += cnt.sum()[0]
				ws[col+str(3+leng)] = sum_
				ws[ncol+str(3+leng)] = str(round(avg_/sum_,3))+'%' if sum_ is not 0 else 'N/A'
				# ORF & NCR
				cnt = bpx[ bpx[book.col_ORF].isin(book.ORF)][[col_diff]]
				ws[col+str(4+leng)] = len(cnt)
				ws[ncol+str(4+leng)] = str(round(cnt.sum()[0] / len(cnt), 3))+'%' if len(cnt) is not 0 else 'N/A'
				cnt = bpx[ bpx[book.col_ORF].isin(book.NCR)][[col_diff]]
				ws[col+str(5+leng)] = len(cnt)
				ws[ncol+str(5+leng)] = str(round(cnt.sum()[0] / len(cnt), 3))+'%' if len(cnt) is not 0 else 'N/A'
				col = next_col(col)

			ncol = chr(ord(ncol)+book.nsheets)
			ws[ncol+str(rows+2)] = str(self.s1[i])+"~"+str(self.s2[i]) if self.s2[i] != 51.0 else str(self.s1[i])+"~"
		infolog("lowhigh GenomeStr end")

	def ORFNCR(self, ws, types_, title, book):
		infolog("lowhigh {0} start".format(title))

		ws['A1'] = title
		ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
		ws['B1'] = 'Length'
		ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

		orfncr = book.ORF if title is 'ORF' else book.NCR
		col_diff = "diffofinc" if types_ == 'INC' else "diffofdec"
		bp = self.BPmergedinc if types_ == 'INC' else self.BPmergeddec

		sum_, col = 0, 'B'
		for si in range(-1, book.nsheets):
			sum_ = 0
			if si != -1:
				ws[col+'2'] = book.sheet_list[si] 
			for r in range(len(orfncr)):
				if si == -1:
					ws['A'+str(r+3)] = orfncr[r]
					cnt = len(book.Dumas[ book.Dumas[book.col_ORF] == orfncr[r] ])
					ws[col+str(r+3)] = cnt
					sum_ += cnt
				else:
					cnt = len( bp[si][ bp[si][book.col_ORF] == orfncr[r] ] )
					ws[col+str(r+3)] = cnt
					sum_ += cnt
			ws[col+str(3+len(orfncr))] = sum_
			col = next_col(col)
		ws['A'+str(3+len(orfncr))] = 'Total'

		col = next_col(col)
		for i in range(len(self.s1)):			
			# ORF or CNR 종류
			ws[col+'1'] = str(self.s1[i])+'~'+str(self.s2[i]) if self.s2[i]!=51.0 else str(self.s1[i])+'~'
			for r in range(len(orfncr)):
				ws[col+str(r+3)] = orfncr[r]
			col = next_col(col)

			col_GPS = col
			ws[col+'1'] = '{0} in genetic polymorphism'.format("Increase" if types_ == 'INC' else "Decrease")
			for s in range(book.nsheets):				
				ws[col+'2'] = book.sheet_list[s]
				bpx = bp[s][ logical_and(bp[s][col_diff] >= self.s1[i], bp[s][col_diff] < self.s2[i]).values ]

				sum_ = 0
				for r in range(3, 3+len(orfncr)):
					cnt = len(bpx[ bpx[book.col_ORF] == orfncr[r-3]])
					ws[col+str(r)] = cnt	
					sum_ += cnt					
				ws[col+str(3+len(orfncr))] = sum_
				col = next_col(col)			

			ws[col+'1'] = 'Average {0}'.format("Increase" if types_ == 'INC' else "Decrease")
			for s in range(book.nsheets):			
				ws[col+'2'] = book.sheet_list[s]
				bpx = bp[s][ logical_and(bp[s][col_diff] >= self.s1[i], bp[s][col_diff] < self.s2[i]).values ]
				sum_ = 0
				for r in range(3, 3+len(orfncr)):
					cnt = bpx[ bpx[book.col_ORF] == orfncr[r-3]][[col_diff]]					
					ws[col+str(r)] = str(round(cnt.sum()[0] / len(cnt), 3))+'%' if len(cnt) is not 0 else 'N/A'
					sum_ += cnt.sum()[0]
				col = next_col(col)

			ws[col+'1'] = 'Number of GPS / Length'
			for s in range(book.nsheets):			
				ws[col+'2'] = book.sheet_list[s]
				for r in range(3, 3+len(orfncr)):
					s, l = float(ws[col_GPS+str(r)].value), float(ws['B'+str(r)].value)
					ws[col+str(r)] = str( round(s/l, 6) ) if l != 0 else 'N/A'
				col_GPS = next_col(col_GPS)
				col = next_col(col)
			col = next_col(col)

		infolog("lowhigh {0} end".format(title))


	def BaseComp(self, ws, types_, book):
		"""
			major가 같은 것, minor도 같은 것만 추출
			minor x -> 증가 포함
			1. major_idx_x == major_idx_y - 치환 제거
			2. (minor_idx_x == minor_idx_y) or ((minor_idx_x != minor_idx_y) and minor_x == 0)
		"""
		infolog("lowhigh BaseComp start")
		ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
		ws['B1'], ws['B2'] = 'Major', 'Minor'
		ws['D1'], ws['H1'], ws['L1'], ws['P1'] = 'A', 'G', 'C', 'T'
		ws['G2'], ws['K2'], ws['O2'] = 'a', 'a', 'a'
		ws['C2'], ws['L2'], ws['P2'] = 'g', 'g', 'g'
		ws['D2'], ws['H2'], ws['Q2'] = 'c', 'c', 'c'
		ws['E2'], ws['I2'], ws['M2'] = 't', 't', 't'
		cols = ['A','G','C','T']		

		bp = self.BPmergedinc if types_ is "INC" else self.BPmergeddec

		for r in range(book.nsheets):
			ws['A'+str(r+3)] = book.sheet_list[r]

		for r in range(book.nsheets):
			col = 'C'
			cond1 = bp[r]['major_idx_x'] == bp[r]['major_idx_y']
			cond2 = bp[r]['minor_idx_x'] == bp[r]['minor_idx_y']
			cond3 = logical_and( bp[r]['minor_idx_x'] != bp[r]['minor_idx_y'], bp[r]['minor_x'] == 0 )

			idx = logical_and(cond1, logical_or(cond2, cond3))
			tmp = bp[r][ idx.values ]

			for a in range(len(cols)):
				for b in range(len(cols)):
					if a == b:
						continue
					# major는 _x, _y 가 이미 같고, minor는 _y기준으로 하면 _x에서 minor가 0이던값 무시됨
					ws[col+str(r+3)] = len(tmp[ logical_and(tmp['major_idx_y'] == a, tmp['minor_idx_y'] == b).values ])
					col = next_col(col)
				col = next_col(col)

		infolog("lowhigh BaseComp end")

	def listofdata(self, ws, n_, types_, book):
<<<<<<< HEAD
		from analyzer import dataframe_to_rows
		print(types_, type(types_))
		# infolog("writing list of {0} data".format(types_))		
		
		bp = self.BPmergedinc[n_] if types_ is "INC" else self.BPmergeddec[n_]
		# bp = bp[ self.dumascol + []]
		rows = dataframe_to_rows(bp, index=False)
=======
		# infolog("writing list of {1} data".format(str(types_)))
		# module import error
		from openpyxl.utils.dataframe import dataframe_to_rows
		bp = self.BPmergedinc[n_] if types_ is "INC" else self.BPmergeddec[n_]
		rows = dataframe_to_rows(bp)

>>>>>>> b1e44ac00ee06e43e40cd7d591d20644bb8c0f16

		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				if str(value) != 'nan':
					ws.cell(row=r_idx, column=c_idx, value=value)

